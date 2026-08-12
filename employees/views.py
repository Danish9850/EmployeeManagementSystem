from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.colors import white,navy, black
from django.http import HttpResponse
from django.db.models import Sum, Avg, Max
from openpyxl import Workbook
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Employee
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from reportlab.lib import colors
from datetime import datetime


# Employee List
@login_required
def employee_list(request):

    search = request.GET.get("search")

    if search:
        employee_queryset = Employee.objects.filter(name__icontains=search)
    else:
        employee_queryset = Employee.objects.all()

    paginator = Paginator(employee_queryset, 5)
    page_number = request.GET.get("page")
    employees = paginator.get_page(page_number)

    total_salary = Employee.objects.aggregate(
        Sum("salary")
    )["salary__sum"] or 0

    average_salary = Employee.objects.aggregate(
        Avg("salary")
    )["salary__avg"] or 0

    highest_salary = Employee.objects.aggregate(
        Max("salary")
    )["salary__max"] or 0

    total_employees = Employee.objects.count()

    return render(request, "employee_list.html", {
        "employees": employees,
        "total_salary": total_salary,
        "average_salary": round(average_salary, 2),
        "highest_salary": highest_salary,
        "total_employees": total_employees,
    })


# Add Employee
@login_required
def add_employee(request):

    if request.method == "POST":

        Employee.objects.create(
            name=request.POST.get("name"),
            email=request.POST.get("email"),
            experience=request.POST.get("experience"),
            salary=request.POST.get("salary"),
            contact=request.POST.get("contact"),
            department=request.POST.get("department"),
            photo=request.FILES.get("photo"),
        )
        messages.success(request, "Employee added successfully.")

        return redirect("employee_list")

    return render(request, "add_employee.html")


# Edit Employee
@login_required
def edit_employee(request, id):

    employee = Employee.objects.get(id=id)

    if request.method == "POST":

        employee.name = request.POST.get("name")
        employee.email = request.POST.get("email")
        employee.experience = request.POST.get("experience")
        employee.salary = request.POST.get("salary")
        employee.contact = request.POST.get("contact")
        employee.department = request.POST.get("department")

        if request.FILES.get("photo"):
            employee.photo = request.FILES.get("photo")

        employee.save()
        messages.success(request, "Employee updated successfully.")

        return redirect("employee_list")

    return render(request, "add_employee.html", {
        "employee": employee
    })


# Delete Employee
@login_required
def delete_employee(request, id):

    employee = get_object_or_404(Employee, id=id)
    if request.method == "POST":
      employee.delete()
      messages.success(request, "Employee deleted successfully.")
      return redirect("employee_list")
    
    return render(request,'delete_confirm.html',{
        'employee':employee
    })

@login_required
def import_employees(request):
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2,values_only=True):
         Employee.objects.create(
             name=row[0],
             email=row[1],
             contact=str(row[2]),
             experience=row[3],
             salary=row[4],
             department=row[5],
         )

        messages.success(request, "Employees imported successfully.")
        return redirect("employee_list")
    return render(request, "import_employees.html")

@login_required
def download_sample(request):
    wb = Workbook()
    ws = wb.active

    ws.append(["Name","Email","Contact","Experience","Salary","Department"])
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment;filename="Employee_Sample.xlsx'
    wb.save(response)
    return response

@login_required
def export_pdf(request):

    response = HttpResponse(content_type="application/pdf")

    response["Content-Disposition"] = 'attachment; filename="Employee_Report.pdf"'

    pdf = canvas.Canvas(response, pagesize=letter)

    width, height = letter

    pdf.setFont("Helvetica-Bold", 20)
    pdf.setFillColor(navy)

    pdf.drawCentredString(
        width/2,
        780,
        "Employee Management System"
    )
    pdf.line(40,770,555,770)

    pdf.setFillColor(black)

    pdf.setFont("Helvetica",11)

    current_date = datetime.now().strftime("%d-%m-%Y %H:%M")
    pdf.drawString(
        40,
        760,
        f"Genrated On : {current_date}"
    )

    pdf.drawString(
        400,
        760,
        f"Generated By : {request.user.username}"
    )

    pdf.setFillColor(black)

    pdf.setFont("Helvetica-Bold",13)

    pdf.drawCentredString(width/2,745,"Employee Report")

    total = Employee.objects.count()

    total_salary = Employee.objects.aggregate(
        Sum("salary")
    )["salary__sum"] or 0

    average_salary = Employee.objects.aggregate(
        Avg("salary")
    )["salary__avg"] or 0

    highest_salary = Employee.objects.aggregate(
        Max("salary")
    )["salary__max"] or 0

    pdf.drawString(40,730,f"Total Employees : {total}")
    pdf.drawString(220,730,f"Total Salary : {total_salary}")

    pdf.drawString(40,710,f"Average Salary : {average_salary:.2f}")
    pdf.drawString(220,710,f"Highest Salary : {highest_salary}")

    y = 670

    pdf.setFillColorRGB(0.15,0.45,0.85)
    pdf.rect(35,y-5,525,22,fill=1)

    pdf.setFillColor(white)
    pdf.setFont("Helvetica-Bold",10)

    pdf.drawString(40,y,"Name")
    pdf.drawString(150,y,"Email")
    pdf.drawString(290,y,"Contact")
    pdf.drawString(390,y,"Salary")
    pdf.drawString(470,y,"Department")
    

    y -= 20

    pdf.setFillColor(black)
    pdf.setFont("Helvetica",10)

    employees = Employee.objects.all()

    for emp in employees:

        pdf.drawString(40,y,emp.name)
        pdf.drawString(150,y,emp.email)
        pdf.drawString(290,y,str(emp.contact))
        pdf.drawString(390,y,str(emp.salary))
        pdf.drawString(470,y,emp.department)
    
        y -= 18

        if y < 50:

            pdf.showPage()

            y = 770

        pdf.setFont("Helvetica-Oblique",9)

        pdf.drawCentredString(width/2,20,"Generated by Employee Management System")
    pdf.save()

    return response